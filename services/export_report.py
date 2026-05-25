"""Exportación PDF y Excel de evaluaciones."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from engine import EvaluationEngine
from services.evaluation_repo import engine_from_evaluation, get_evaluation_detail


def _build_rows(engine: EvaluationEngine) -> list[dict]:
    rows = []
    for i, name in enumerate(engine.factor_names):
        rel_label, _ = engine.relative_importance(i)
        st = engine.factor_states[i]
        status = engine.foda_row_status(i)
        rows.append(
            {
                "Factor": name,
                "Importancia relativa": rel_label,
                "Ponderación": f"{st.global_weight:.1f}" if st.global_weight else "—",
                "Alcance": st.scope,
                "FODA": st.foda if status == "done" else (
                    "Pendiente" if status == "pending" else "No relevante"
                ),
            }
        )
    return rows


def export_excel(evaluacion_id: int) -> bytes:
    engine, _ = engine_from_evaluation(evaluacion_id)
    meta = get_evaluation_detail(evaluacion_id) or {}
    rows = _build_rows(engine)
    text, style = engine.compute_recommendation()

    wb = Workbook()
    ws = wb.active
    ws.title = "Informe GUIOSAD"

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)

    ws["A1"] = "Informe de evaluación FLOSS — GUIOSAD"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Proyecto: {meta.get('nombre_proyecto', '')}"
    ws["A3"] = f"Software: {meta.get('software_nombre', '')}"
    ws["A4"] = f"Re-evaluación #: {meta.get('numero_reevaluacion', 1)}"
    ws["A5"] = f"Estado: {meta.get('estado', '')}"
    ws["A6"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    start = 8
    headers = list(rows[0].keys()) if rows else []
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, start + 1):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row[h])

    rec_row = start + len(rows) + 2
    ws.cell(row=rec_row, column=1, value="Recomendación final").font = Font(bold=True)
    ws.cell(row=rec_row + 1, column=1, value=text or meta.get("recomendacion_texto", ""))
    ws.merge_cells(start_row=rec_row + 1, start_column=1, end_row=rec_row + 3, end_column=len(headers) or 5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(evaluacion_id: int) -> bytes:
    engine, _ = engine_from_evaluation(evaluacion_id)
    meta = get_evaluation_detail(evaluacion_id) or {}
    rows = _build_rows(engine)
    text, _ = engine.compute_recommendation()
    rec = text or meta.get("recomendacion_texto", "Sin recomendación generada.")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2 * cm, leftMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], textColor=colors.HexColor("#0f766e"))
    story = [
        Paragraph("Informe GUIOSAD — Adopción FLOSS", title_style),
        Spacer(1, 12),
        Paragraph(f"<b>Proyecto:</b> {meta.get('nombre_proyecto', '')}"),
        Paragraph(f"<b>Software:</b> {meta.get('software_nombre', '')}"),
        Paragraph(f"<b>Re-evaluación #:</b> {meta.get('numero_reevaluacion', 1)} | <b>Estado:</b> {meta.get('estado', '')}"),
        Paragraph(f"<b>Fecha informe:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}"),
        Spacer(1, 16),
    ]

    if rows:
        data = [list(rows[0].keys())] + [[str(row[k]) for k in rows[0].keys()] for row in rows]
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        )
        story.append(table)

    story.extend(
        [
            Spacer(1, 20),
            Paragraph("<b>Recomendación final</b>", styles["Heading2"]),
            Paragraph(rec.replace("\n", "<br/>"), styles["Normal"]),
        ]
    )
    doc.build(story)
    return buf.getvalue()
